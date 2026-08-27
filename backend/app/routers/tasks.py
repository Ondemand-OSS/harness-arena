from __future__ import annotations

import datetime as dt
import io
import json
import os
import tempfile

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pymongo.database import Database
from starlette.responses import Response

from ..cache import get_json as cache_get, invalidate as cache_invalidate, mark_response as cache_mark, set_json as cache_set
from ..dataset_import import EXPECTED_HEADERS, import_xlsx
from ..db import get_db, next_id
from ..logger import log_activity
from ..runner import _guess_media_type
from ..schemas import CategoryReviewApproveIn, ImportResult, ReferenceFileOut, TaskOut
from ..taxonomy import (
    DEFAULT_GROUP,
    GROUPS,
    category_key,
    deliverable_types,
    group_for_category_with_approvals,
    is_builtin_category,
    parse_deliverables,
)
from ..users import current_user, is_admin, require_arena_admin, require_user

router = APIRouter(prefix="/api/tasks", tags=["tasks"])

# Ceiling on `list_tasks`'s `limit` query param — keeps one request from
# demanding the whole (unbounded) task set back as a single "page", the
# same way runs.py's admin_list_runs caps its own `limit`.
MAX_TASKS_PAGE_LIMIT = 50


def _approved_category_groups(db: Database) -> dict[str, str]:
    return {
        review["_id"]: review.get("group") if review.get("group") in {*GROUPS, DEFAULT_GROUP} else DEFAULT_GROUP
        for review in db.category_reviews.find({"status": "approved"}, {"group": 1})
    }


def task_out(
    doc: dict,
    db: Database,
    has_judge_verdict: bool = False,
    approved_groups: dict[str, str] | None = None,
    submitter: dict | None = None,
) -> TaskOut:
    """Serialize a task document plus its derived taxonomy/deliverable
    fields and its submitter's display info."""
    if submitter is None and doc.get("submitted_by_user_id") is not None:
        submitter = db.users.find_one({"_id": doc["submitted_by_user_id"]})
    return TaskOut(
        has_judge_verdict=has_judge_verdict,
        submitted_by=(submitter.get("display_name") or submitter.get("username")) if submitter else None,
        submitted_by_avatar=submitter.get("avatar_key", "") if submitter else "",
        is_deleted=bool(doc.get("is_deleted")),
        results_deleted=bool(doc.get("results_deleted")),
        imported_at=doc.get("imported_at"),
        id_aa=doc["_id"],
        title=doc.get("title", ""),
        category=doc.get("category", ""),
        prompt=doc.get("prompt", ""),
        system_prompt=doc.get("system_prompt", ""),
        rubric=doc.get("rubric", ""),
        expected_deliverables=doc.get("expected_deliverables", ""),
        reference_files=doc.get("reference_files", ""),
        dataset_version=doc.get("dataset_version", ""),
        group=group_for_category_with_approvals(doc.get("category", ""), approved_groups),
        deliverable_files=parse_deliverables(doc.get("expected_deliverables", "")),
        deliverable_types=deliverable_types(doc.get("expected_deliverables", "")),
    )


_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
_DATASET_FILENAME = "Multi_Source_Agent_Workflows-dataset.xlsx"
# `extras/` is the local-only, gitignored home for real bench inputs (the
# dataset, real deliverables, judge results — see scripts/import_seed_results.py).
# The repo root is checked too for anyone who keeps the dataset there instead.
DEFAULT_DATASET_PATH = next(
    (
        p
        for p in (
            os.path.join(_REPO_ROOT, "extras", _DATASET_FILENAME),
            os.path.join(_REPO_ROOT, _DATASET_FILENAME),
        )
        if os.path.isfile(p)
    ),
    os.path.join(_REPO_ROOT, _DATASET_FILENAME),
)


@router.get("", response_model=list[TaskOut])
def list_tasks(response: Response, category: str | None = None, group: str | None = None, include_deleted: bool = False, page: int = 1, limit: int | None = None, db: Database = Depends(get_db), user: dict | None = Depends(current_user)):
    if include_deleted and not is_admin(user):
        raise HTTPException(status_code=403, detail="only the OnDemand admin can view deleted tasks")
    # `group` is derived (see below), not a stored field, so pagination has
    # to happen in-memory AFTER that filter — a Mongo-level skip/limit here
    # would slice before group filtering and hand back short, misaligned
    # pages. `limit` is opt-in: omitting it (the default) returns every
    # task exactly as before, for every existing caller. Clamp up front so
    # the cache key reflects the request's actual, effective page/limit.
    if limit is not None:
        page = max(1, page)
        limit = max(1, min(limit, MAX_TASKS_PAGE_LIMIT))
    cache_variant = json.dumps([category or "", group or "", page if limit is not None else None, limit], separators=(",", ":"))
    if not include_deleted:
        cached = cache_get("tasks", f"list:{cache_variant}")
        if cached is not None:
            cache_mark(response, hit=True)
            return cached
    cache_mark(response, hit=False)
    query = {"category": category} if category else {}
    if not include_deleted:
        query["is_deleted"] = {"$ne": True}
    tasks = list(db.tasks.find(query).sort("_id", 1))
    # One query for the whole page rather than a per-task lookup.
    with_verdicts = {v["task_id"] for v in db.judge_verdicts.find({"is_deleted": {"$ne": True}}, {"task_id": 1})}
    approved_groups = _approved_category_groups(db)
    submitter_ids = {t["submitted_by_user_id"] for t in tasks if t.get("submitted_by_user_id") is not None}
    submitters = {u["_id"]: u for u in db.users.find({"_id": {"$in": list(submitter_ids)}})} if submitter_ids else {}
    out = [
        task_out(
            t,
            db,
            has_judge_verdict=t["_id"] in with_verdicts,
            approved_groups=approved_groups,
            submitter=submitters.get(t.get("submitted_by_user_id"), {}),
        )
        for t in tasks
    ]
    # `group` is derived, not a stored field, so it has to be filtered after
    # serialization rather than in the Mongo query.
    if group:
        out = [t for t in out if t.group == group]
    if limit is not None:
        start = (page - 1) * limit
        out = out[start : start + limit]
    if not include_deleted:
        cache_set("tasks", out, variant=f"list:{cache_variant}", ttl_seconds=90)
    return out


@router.get("/categories", response_model=list[str])
def list_categories(response: Response, group: str | None = None, db: Database = Depends(get_db)):
    """Distinct categories, optionally narrowed to one coarse group — so a
    category filter never offers options that belong to a different group
    and would return nothing."""
    cache_variant = f"categories:{group or ''}"
    cached = cache_get("tasks", cache_variant)
    if cached is not None:
        cache_mark(response, hit=True)
        return cached
    cache_mark(response, hit=False)
    approved_groups = _approved_category_groups(db)
    cats = {
        c
        for c in db.tasks.distinct("category", {"is_deleted": {"$ne": True}})
        if c and (is_builtin_category(c) or category_key(c) in approved_groups)
    }
    if group:
        cats = {c for c in cats if group_for_category_with_approvals(c, approved_groups) == group}
    out = sorted(cats)
    cache_set("tasks", out, variant=cache_variant, ttl_seconds=120)
    return out


@router.get("/groups", response_model=list[str])
def list_groups(response: Response, db: Database = Depends(get_db)):
    """Coarse groups that actually have tasks, in canonical GROUPS order —
    so the top-bar nav never shows an empty group."""
    cached = cache_get("tasks", "groups")
    if cached is not None:
        cache_mark(response, hit=True)
        return cached
    cache_mark(response, hit=False)
    approved_groups = _approved_category_groups(db)
    present = {
        group_for_category_with_approvals(c, approved_groups)
        for c in db.tasks.distinct("category", {"is_deleted": {"$ne": True}})
        if c
    }
    ordered = [g for g in GROUPS if g in present]
    if DEFAULT_GROUP in present:
        ordered.append(DEFAULT_GROUP)
    cache_set("tasks", ordered, variant="groups", ttl_seconds=120)
    return ordered


def _ensure_pending_category_reviews(db: Database) -> None:
    """Backfill review records for categories uploaded before this feature."""
    now = dt.datetime.now(dt.timezone.utc)
    for category in db.tasks.distinct("category", {"is_deleted": {"$ne": True}}):
        category = (category or "").strip()
        if category and not is_builtin_category(category):
            db.category_reviews.update_one(
                {"_id": category_key(category)},
                {"$setOnInsert": {"category": category, "status": "pending", "created_at": now}, "$set": {"last_seen_at": now}},
                upsert=True,
            )


@router.get("/category-reviews")
def list_category_reviews(db: Database = Depends(get_db), _admin: dict = Depends(require_arena_admin)):
    """Uploaded categories an admin can approve or place in a group."""
    _ensure_pending_category_reviews(db)
    reviews = list(db.category_reviews.find({"status": {"$in": ["pending", "approved"]}}).sort("category", 1))
    return [
        {
            "key": review["_id"],
            "category": review.get("category", review["_id"]),
            "task_count": db.tasks.count_documents({"category": review.get("category", "") , "is_deleted": {"$ne": True}}),
            "status": review.get("status", "pending"),
            "group": review.get("group", ""),
        }
        for review in reviews
    ]


@router.put("/category-reviews/{review_key:path}/approve")
def approve_category_review(
    review_key: str,
    body: CategoryReviewApproveIn,
    db: Database = Depends(get_db),
    admin: dict = Depends(require_arena_admin),
):
    review = db.category_reviews.find_one({"_id": review_key})
    if review is None:
        raise HTTPException(status_code=404, detail="category review not found")
    if body.group not in {*GROUPS, DEFAULT_GROUP}:
        raise HTTPException(status_code=400, detail="invalid category group")
    db.category_reviews.update_one(
        {"_id": review_key},
        {"$set": {"status": "approved", "group": body.group, "approved_at": dt.datetime.now(dt.timezone.utc)}},
    )
    cache_invalidate("tasks")
    log_activity(
        db,
        action="CATEGORY_REVIEW_APPROVE",
        user_id=admin["_id"],
        message=f"approved category review {review_key} in {body.group}",
        metadata={"review_key": review_key, "category": review.get("category"), "group": body.group},
        route="/api/tasks/category-reviews/{review_key}/approve",
    )
    return {"category": review.get("category", review_key), "group": body.group}


@router.get("/template.xlsx")
def download_dataset_template():
    """A fill-in Excel workbook for custom benchmark tasks.

    It intentionally contains no task rows: a downloaded-but-unedited
    template should not quietly import a demonstration task into a user's
    benchmark.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = openpyxl.Workbook()
    tasks_sheet = workbook.active
    tasks_sheet.title = "Tasks"
    tasks_sheet.append(EXPECTED_HEADERS)
    tasks_sheet.freeze_panes = "A2"
    tasks_sheet.auto_filter.ref = "A1:H1"
    header_fill = PatternFill("solid", fgColor="1F4D3B")
    for cell in tasks_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column, width in zip("ABCDEFGH", (24, 32, 24, 60, 48, 52, 40, 40)):
        tasks_sheet.column_dimensions[column].width = width

    instructions = workbook.create_sheet("Instructions")
    instructions.column_dimensions["A"].width = 27
    instructions.column_dimensions["B"].width = 100
    instructions.append(["Custom benchmark dataset", "Fill in one task per row on the Tasks sheet, then upload the .xlsx file from New benchmark."])
    instructions.append(["Column", "What to provide"])
    guidance = [
        ("id_aa", "Required. A unique, stable task ID, e.g. my_benchmark_001. Re-uploading the same ID updates that task."),
        ("title", "A short, human-readable task name."),
        ("category", "Any category label. Known categories are grouped automatically; unfamiliar labels appear under Other."),
        ("prompt", "The complete task instruction given to each agent."),
        ("system_prompt", "Optional additional agent instructions. Leave blank when you do not need them."),
        ("rubric", "Optional judging criteria for the task."),
        (
            "expected_deliverables",
            "Optional comma-separated exact filenames, e.g. analysis.xlsx, summary.md. Maximum 20 filenames per task.",
        ),
        (
            "reference_files",
            "Write na if this task needs no reference material. Otherwise, the exact filename(s) of reference "
            "material this task's prompt names (comma-separated for more than one), e.g. "
            "14_digital_spirituality_ecosystem.md — currently .md only. Naming a file here only describes it: "
            "you must also upload its actual content from the task's row in Benchmark (or POST "
            "/api/tasks/{id}/reference-files) before this task can be submitted — a task that names a file "
            "with nothing uploaded for it blocks the whole benchmark from starting until it's attached.",
        ),
    ]
    for row in guidance:
        instructions.append(row)
    instructions.freeze_panes = "A3"
    for cell in instructions[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for cell in instructions[2]:
        cell.font = Font(bold=True)
    for row in instructions.iter_rows(min_row=1, max_row=instructions.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    content = io.BytesIO()
    workbook.save(content)
    return Response(
        content=content.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="arena-custom-dataset-template.xlsx"'},
    )


def _ensure_not_running(db: Database, task_id: str) -> None:
    """Avoid a background batch recreating data while it is being removed.

    Scoped to `current_task_id == task_id` — the one task a running batch
    is actively working on right this moment (see batches.py's
    _execute_batch) — not just "this task_id appears somewhere in a batch
    that's still running overall". The old, broader check blocked deleting
    ANY task that was ever part of a still-in-progress multi-task batch,
    including one the batch had already finished with (or hadn't reached
    yet), for as long as the rest of the batch kept running — sometimes
    indefinitely, if that batch itself got orphaned by a server restart
    (see runner.reconcile_orphaned_batches, which now resolves that case
    directly instead of leaving it stuck "running" forever)."""
    if db.batches.find_one({"status": "running", "current_task_id": task_id}, {"_id": 1}):
        raise HTTPException(status_code=409, detail="this task is currently running; wait for it to finish before deleting it")


def _delete_results(db: Database, task_id: str) -> None:
    archived_at = dt.datetime.now(dt.timezone.utc)
    db.runs.update_many({"task_id": task_id}, {"$set": {"is_deleted": True, "deleted_at": archived_at}})
    db.scores.update_many({"task_id": task_id}, {"$set": {"is_deleted": True, "deleted_at": archived_at}})
    db.judge_verdicts.update_many({"task_id": task_id}, {"$set": {"is_deleted": True, "deleted_at": archived_at}})
    db.tasks.update_one({"_id": task_id}, {"$set": {"results_deleted": True, "results_deleted_at": archived_at}})


@router.delete("/{task_id}/results", status_code=204)
def delete_task_results(task_id: str, db: Database = Depends(get_db), admin: dict = Depends(require_arena_admin)):
    if db.tasks.find_one({"_id": task_id}, {"_id": 1}) is None:
        raise HTTPException(status_code=404, detail="task not found")
    _ensure_not_running(db, task_id)
    _delete_results(db, task_id)
    cache_invalidate("tasks", "stats", "leaderboard")
    log_activity(
        db,
        action="TASK_RESULTS_DELETE",
        user_id=admin["_id"],
        message=f"soft-deleted every run, score and verdict for task {task_id}",
        metadata={"task_id": task_id},
        route="/api/tasks/{task_id}/results",
    )


@router.delete("/{task_id}", status_code=204)
def delete_task(task_id: str, db: Database = Depends(get_db), admin: dict = Depends(require_arena_admin)):
    if db.tasks.find_one({"_id": task_id}, {"_id": 1}) is None:
        raise HTTPException(status_code=404, detail="task not found")
    _ensure_not_running(db, task_id)
    _delete_results(db, task_id)
    db.tasks.update_one({"_id": task_id}, {"$set": {"is_deleted": True}})
    cache_invalidate("tasks", "stats", "leaderboard")
    log_activity(
        db,
        action="TASK_DELETE",
        user_id=admin["_id"],
        message=f"soft-deleted task {task_id} and its results",
        metadata={"task_id": task_id},
        route="/api/tasks/{task_id}",
    )


@router.post("/{task_id}/restore", status_code=204)
def restore_task(task_id: str, db: Database = Depends(get_db), admin: dict = Depends(require_arena_admin)):
    task = db.tasks.find_one({"_id": task_id})
    if task is None:
        raise HTTPException(status_code=404, detail="task not found")
    db.tasks.update_one(
        {"_id": task_id},
        {"$set": {"is_deleted": False, "results_deleted": False}, "$unset": {"deleted_at": "", "results_deleted_at": ""}},
    )
    db.runs.update_many({"task_id": task_id}, {"$set": {"is_deleted": False}, "$unset": {"deleted_at": ""}})
    db.scores.update_many({"task_id": task_id}, {"$set": {"is_deleted": False}, "$unset": {"deleted_at": ""}})
    db.judge_verdicts.update_many({"task_id": task_id}, {"$set": {"is_deleted": False}, "$unset": {"deleted_at": ""}})
    cache_invalidate("tasks", "stats", "leaderboard")
    log_activity(
        db,
        action="TASK_RESTORE",
        user_id=admin["_id"],
        message=f"restored task {task_id} and its results",
        metadata={"task_id": task_id},
        route="/api/tasks/{task_id}/restore",
    )


@router.get("/{task_id}", response_model=TaskOut)
def get_task(task_id: str, response: Response, db: Database = Depends(get_db)):
    cached = cache_get("tasks", f"task:{task_id}")
    if cached is not None:
        cache_mark(response, hit=True)
        return cached
    cache_mark(response, hit=False)
    task = db.tasks.find_one({"_id": task_id})
    if task is None:
        raise HTTPException(status_code=404, detail="task not found")
    has_verdict = db.judge_verdicts.find_one({"task_id": task_id}) is not None
    out = task_out(task, db, has_judge_verdict=has_verdict, approved_groups=_approved_category_groups(db))
    cache_set("tasks", out, variant=f"task:{task_id}", ttl_seconds=90)
    return out


# Every reference file seen in this dataset so far is a Markdown scope/
# sensitivity brief (see extras/reference_files/) — restricted to that for
# now rather than accepting anything, so a stray upload of the wrong file
# type fails loudly at upload time instead of silently confusing a run
# later. Relax this (e.g. add .txt/.pdf) once a real task needs it.
# .json added for the Terminal Agent Harness dataset's API-delivery tasks
# (software_api_delivery_009/010), whose reference material is a JSON
# fixture (application_inputs.json) rather than a Markdown brief.
_ALLOWED_REFERENCE_FILE_EXTENSIONS = {".md", ".json"}
_MAX_REFERENCE_FILE_BYTES = 20 * 1024 * 1024


def _reference_file_out(doc: dict) -> ReferenceFileOut:
    return ReferenceFileOut(
        id=doc["_id"],
        task_id=doc["task_id"],
        filename=doc["filename"],
        media_type=doc["media_type"],
        size_bytes=doc["size_bytes"],
        uploaded_at=doc.get("uploaded_at"),
    )


@router.get("/{task_id}/reference-files", response_model=list[ReferenceFileOut])
def list_reference_files(task_id: str, db: Database = Depends(get_db)):
    """Metadata only (no bytes) — what `reference_files` names as text can
    now optionally have real content attached via the endpoints below; this
    lists what's actually attached, for the task detail / setup UI."""
    if db.tasks.find_one({"_id": task_id}, {"_id": 1}) is None:
        raise HTTPException(status_code=404, detail="task not found")
    docs = db.task_reference_files.find({"task_id": task_id}).sort("_id", 1)
    return [_reference_file_out(d) for d in docs]


@router.post("/{task_id}/reference-files", response_model=ReferenceFileOut)
async def upload_reference_file(
    task_id: str,
    file: UploadFile = File(...),
    db: Database = Depends(get_db),
    admin: dict = Depends(require_arena_admin),
):
    """Attaches a real reference file's bytes to a task, so a run can
    actually hand the material to the agent instead of only naming it in
    `reference_files` text (see harnesses/_prompt.py and
    harnesses/_reference_files.py for how a run picks these up). Re-
    uploading the same filename replaces the previous content rather than
    creating a duplicate."""
    if db.tasks.find_one({"_id": task_id}, {"_id": 1}) is None:
        raise HTTPException(status_code=404, detail="task not found")
    filename = os.path.basename(file.filename or "").strip()
    if not filename:
        raise HTTPException(status_code=400, detail="the uploaded file needs a filename")
    ext = os.path.splitext(filename)[1].lower()
    if ext not in _ALLOWED_REFERENCE_FILE_EXTENSIONS:
        allowed = ", ".join(sorted(_ALLOWED_REFERENCE_FILE_EXTENSIONS))
        raise HTTPException(status_code=400, detail=f"reference files must be one of: {allowed}")
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="the uploaded file is empty")
    if len(content) > _MAX_REFERENCE_FILE_BYTES:
        raise HTTPException(
            status_code=400,
            detail=f"reference files are capped at {_MAX_REFERENCE_FILE_BYTES // (1024 * 1024)} MB",
        )
    doc = {
        "task_id": task_id,
        "filename": filename,
        "media_type": _guess_media_type(filename),
        "size_bytes": len(content),
        "content": content,
        "uploaded_at": dt.datetime.now(dt.timezone.utc),
    }
    existing = db.task_reference_files.find_one({"task_id": task_id, "filename": filename}, {"_id": 1})
    if existing:
        db.task_reference_files.update_one({"_id": existing["_id"]}, {"$set": doc})
        doc["_id"] = existing["_id"]
    else:
        doc["_id"] = next_id(db, "task_reference_files")
        db.task_reference_files.insert_one(doc)
    # The file's bytes are deliberately not in the log — only what it was
    # and how big, which is what a "who attached what, when" trail needs.
    log_activity(
        db,
        action="REFERENCE_FILE_UPLOAD",
        user_id=admin["_id"],
        message=f"{'replaced' if existing else 'attached'} reference file {filename} on task {task_id}",
        metadata={
            "task_id": task_id,
            "file_id": doc["_id"],
            "filename": filename,
            "media_type": doc["media_type"],
            "size_bytes": doc["size_bytes"],
            "replaced_existing": bool(existing),
        },
        route="/api/tasks/{task_id}/reference-files",
    )
    return _reference_file_out(doc)


@router.delete("/{task_id}/reference-files/{file_id}", status_code=204)
def delete_reference_file(
    task_id: str, file_id: int, db: Database = Depends(get_db), admin: dict = Depends(require_arena_admin)
):
    doc = db.task_reference_files.find_one({"_id": file_id, "task_id": task_id}, {"content": 0})
    result = db.task_reference_files.delete_one({"_id": file_id, "task_id": task_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="reference file not found")
    log_activity(
        db,
        action="REFERENCE_FILE_DELETE",
        user_id=admin["_id"],
        message=f"deleted reference file {(doc or {}).get('filename', file_id)} from task {task_id}",
        metadata={"task_id": task_id, "file_id": file_id, "filename": (doc or {}).get("filename"), "size_bytes": (doc or {}).get("size_bytes")},
        route="/api/tasks/{task_id}/reference-files/{file_id}",
    )


@router.get("/{task_id}/reference-files/{file_id}/content")
def get_reference_file_content(
    task_id: str, file_id: int, db: Database = Depends(get_db), _user: dict = Depends(require_user)
):
    """Return reference material to a signed-in judge.

    These files are part of the instructions given to every harness, so they
    must also be available to a judge assessing the resulting deliverables.
    Upload and deletion remain arena-admin-only.
    """
    doc = db.task_reference_files.find_one({"_id": file_id, "task_id": task_id})
    if doc is None:
        raise HTTPException(status_code=404, detail="reference file not found")
    return Response(
        content=doc["content"],
        media_type=doc["media_type"],
        headers={"Content-Disposition": f'attachment; filename="{os.path.basename(doc["filename"])}"'},
    )


def _latest_dataset_version(db: Database) -> str:
    row = db.tasks.find_one(sort=[("imported_at", -1)])
    return row.get("dataset_version", "") if row else ""


@router.post("/import-default", response_model=ImportResult)
def import_default_dataset(db: Database = Depends(get_db), user: dict | None = Depends(current_user)):
    """Import the bundled Multi_Source_Agent_Workflows-dataset.xlsx that
    ships in the repo root."""
    if not os.path.isfile(DEFAULT_DATASET_PATH):
        raise HTTPException(status_code=404, detail="bundled dataset not found on disk")
    try:
        count, new_task_ids, skipped_existing_ids = import_xlsx(DEFAULT_DATASET_PATH, db)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    cache_invalidate("tasks", "stats", "leaderboard")
    # `current_user` is for attribution only — this endpoint has never
    # required auth, so it stays callable unauthenticated (user_id None).
    log_activity(
        db,
        action="DATASET_IMPORT_DEFAULT",
        user_id=user["_id"] if user else None,
        message=f"imported the bundled dataset ({count} new task(s), {len(skipped_existing_ids)} already existed and were left unchanged)",
        metadata={
            "imported": count,
            "new_task_ids": new_task_ids,
            "skipped_existing_ids": skipped_existing_ids,
            "dataset_version": _latest_dataset_version(db),
        },
        route="/api/tasks/import-default",
    )
    return ImportResult(
        imported=count,
        dataset_version=_latest_dataset_version(db),
        new_task_ids=new_task_ids,
        skipped_existing_ids=skipped_existing_ids,
    )


@router.post("/import", response_model=ImportResult)
async def import_dataset(file: UploadFile = File(...), db: Database = Depends(get_db), user: dict | None = Depends(current_user)):
    """Upload a new/updated xlsx dataset (same "Tasks" sheet schema). Any
    file that isn't a valid custom or Artificial Analysis workbook — wrong format,
    missing columns, nothing importable — is rejected with a clear reason
    (see dataset_import.import_xlsx) rather than partially imported."""
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=400,
            detail="This doesn't look like a valid dataset. Expected an .xlsx file.",
        )
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name
    try:
        count, new_task_ids, skipped_existing_ids = import_xlsx(tmp_path, db)
    except ValueError as exc:
        # A rejected upload is exactly the kind of thing someone reports as
        # "the import didn't work" — keep the reason and the filename.
        log_activity(
            db,
            action="DATASET_IMPORT",
            status="FAILED",
            user_id=user["_id"] if user else None,
            message=f"dataset upload rejected: {exc}",
            metadata={"filename": file.filename},
            route="/api/tasks/import",
        )
        raise HTTPException(status_code=400, detail=str(exc))
    finally:
        os.unlink(tmp_path)
    cache_invalidate("tasks", "stats", "leaderboard")
    log_activity(
        db,
        action="DATASET_IMPORT",
        user_id=user["_id"] if user else None,
        message=(
            f"imported dataset {file.filename} ({count} new task(s), "
            f"{len(skipped_existing_ids)} already existed and were left unchanged)"
        ),
        metadata={
            "filename": file.filename,
            "imported": count,
            "new_task_ids": new_task_ids,
            "skipped_existing_ids": skipped_existing_ids,
            "dataset_version": _latest_dataset_version(db),
        },
        route="/api/tasks/import",
    )
    return ImportResult(
        imported=count,
        dataset_version=_latest_dataset_version(db),
        new_task_ids=new_task_ids,
        skipped_existing_ids=skipped_existing_ids,
    )
