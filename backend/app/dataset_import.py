"""Import tasks from an XLSX dataset."""
from __future__ import annotations

import datetime as dt

import openpyxl
from pymongo.database import Database

from .taxonomy import category_key, is_builtin_category

EXPECTED_HEADERS = [
    "id_aa",
    "title",
    "category",
    "prompt",
    "system_prompt",
    "rubric",
    "expected_deliverables",
    "reference_files",
]
MAX_EXPECTED_DELIVERABLES_PER_TASK = 20


def _deliverable_count(value) -> int:
    if value is None:
        return 0
    return sum(1 for name in str(value).split(",") if name.strip())


def import_xlsx(path: str, db: Database, dataset_version: str | None = None) -> tuple[int, list[str], list[str]]:
    """Import a workbook without overwriting existing task records."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(
            f"This doesn't look like a valid dataset. The file couldn't be read as an "
            f".xlsx workbook ({type(exc).__name__}). Please check the export and try again."
        ) from exc

    sheet = wb["Tasks"] if "Tasks" in wb.sheetnames else wb.worksheets[0]

    rows = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        raise ValueError("This doesn't look like a valid dataset. The sheet is empty.")

    header = [str(h).strip() if h is not None else "" for h in header_row]
    col_index = {name: header.index(name) for name in EXPECTED_HEADERS if name in header}
    missing = [h for h in EXPECTED_HEADERS if h not in col_index]
    if missing:
        raise ValueError(
            "This doesn't look like a valid dataset. Missing required column(s): "
            f"{', '.join(missing)}. Expected: {', '.join(EXPECTED_HEADERS)}."
        )

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row is None or all(cell is None for cell in row):
            continue
        id_value = row[col_index["id_aa"]] if col_index["id_aa"] < len(row) else None
        if id_value is None or str(id_value).strip() == "":
            continue
        expected_value = row[col_index["expected_deliverables"]] if col_index["expected_deliverables"] < len(row) else None
        count = _deliverable_count(expected_value)
        if count > MAX_EXPECTED_DELIVERABLES_PER_TASK:
            raise ValueError(
                f"Task '{str(id_value).strip()}' has {count} expected deliverables. "
                f"Maximum {MAX_EXPECTED_DELIVERABLES_PER_TASK} deliverables allowed per task."
            )

    version = dataset_version or dt.datetime.now(dt.timezone.utc).strftime("v%Y%m%d%H%M%S")

    count = 0
    new_task_ids: list[str] = []
    skipped_existing_ids: list[str] = []
    for row in rows:
        if row is None or all(cell is None for cell in row):
            continue
        id_aa = row[col_index["id_aa"]]
        if id_aa is None or str(id_aa).strip() == "":
            continue
        id_aa = str(id_aa).strip()

        if db.tasks.find_one({"_id": id_aa}, {"_id": 1}) is not None:
            skipped_existing_ids.append(id_aa)
            continue

        def cell(name: str) -> str:
            idx = col_index[name]
            val = row[idx] if idx < len(row) else None
            return "" if val is None else str(val)

        doc = {
            "_id": id_aa,
            "id_aa": id_aa,
            "title": cell("title"),
            "category": cell("category"),
            "prompt": cell("prompt"),
            "system_prompt": cell("system_prompt"),
            "rubric": cell("rubric"),
            "expected_deliverables": cell("expected_deliverables"),
            "reference_files": cell("reference_files"),
            "dataset_version": version,
            "is_deleted": False,
            "imported_at": dt.datetime.now(dt.timezone.utc),
            "submitted_by_user_id": None,
        }
        db.tasks.insert_one(doc)
        new_task_ids.append(id_aa)
        category = doc["category"].strip()
        if category and not is_builtin_category(category):
            db.category_reviews.update_one(
                {"_id": category_key(category)},
                {
                    "$setOnInsert": {
                        "category": category,
                        "status": "pending",
                        "created_at": dt.datetime.now(dt.timezone.utc),
                    },
                    "$set": {"last_seen_at": dt.datetime.now(dt.timezone.utc)},
                },
                upsert=True,
            )
        count += 1

    if count == 0 and not skipped_existing_ids:
        raise ValueError(
            "This doesn't look like a valid dataset. The required columns are present, but "
            "no row had a usable id_aa value, so zero tasks could be imported."
        )
    return count, new_task_ids, skipped_existing_ids
