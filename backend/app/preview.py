"""Turns a deliverable file's bytes into structured JSON the browser can
render inline, so a judge can actually read the work instead of downloading
five office files per output.

Takes bytes rather than a file path: deliverable content lives in the
`deliverables` collection as BSON Binary (see runner.py), not on local
disk, so every extractor here reads from an in-memory BytesIO instead of
opening a path. That also happens to be exactly what each underlying
library (openpyxl, python-docx, python-pptx, pypdf) already accepts.

Every extractor is defensive: these are files produced by an agent, not
trusted input, and a malformed one must degrade to an "unavailable" preview
rather than 500 the judging page. Output is also size-capped in every
dimension (sheets, rows, columns, characters, slides, pages) — a judge
needs enough to assess the work, not a faithful reproduction of a 4MB PDF.
"""
from __future__ import annotations

import io
import json
import os
import subprocess
import tempfile
from pathlib import Path

MAX_SHEETS = 40
MAX_ROWS = 5000
MAX_COLS = 200
MAX_CHARS = 60_000
MAX_SLIDES = 60
MAX_PDF_PAGES = 40
MAX_CELL_CHARS = 500
PPTX_RENDER_TIMEOUT_SECONDS = 60
MAX_RENDERED_PDF_BYTES = 100 * 1024 * 1024


def _clip(text: str, limit: int = MAX_CHARS) -> tuple[str, bool]:
    if text is None:
        return "", False
    if len(text) <= limit:
        return text, False
    return text[:limit], True


def _sheet_cells(row) -> list[str]:
    cells = []
    for value in row[:MAX_COLS]:
        if value is None:
            cells.append("")
        else:
            text = str(value)
            cells.append(text if len(text) <= MAX_CELL_CHARS else text[:MAX_CELL_CHARS] + "…")
    return cells


def _read_sheet(ws) -> tuple[list[list[str]], bool, bool]:
    """Reads one worksheet. Returns (rows, unreadable, more_rows).

    Iterates the row generator by hand (rather than a plain `for`) so a bad
    cell partway through a sheet — e.g. a shared-string index with no matching
    entry, or a numeric cell carrying non-numeric text, both of which openpyxl
    resolves lazily per row in read_only mode and raises on — can be caught
    without losing the rows already read. Once the generator raises it's dead
    (Python generators can't resume past a raised exception), so we salvage
    what we have and move on instead of failing the whole file.

    Blank rows are SKIPPED rather than counted against MAX_ROWS. That is the
    difference between reading a sheet and reporting it empty: read_only mode
    trusts the worksheet's declared <dimension>, so a file with a bogus range
    or a leading blank block yields blank rows first, and the old code spent
    its entire row budget on them, popped the trailing blanks, and returned
    nothing at all for a sheet that plainly had data in it.
    """
    rows: list[list[str]] = []
    unreadable = False
    more_rows = False
    row_iter = iter(ws.iter_rows(values_only=True))
    while True:
        try:
            row = next(row_iter)
        except StopIteration:
            break
        except Exception:
            unreadable = True
            break
        cells = _sheet_cells(row)
        if not any(c != "" for c in cells):
            continue
        if len(rows) >= MAX_ROWS:
            more_rows = True
            break
        rows.append(cells)
    return rows, unreadable, more_rows


def _xlsx_from_workbook(wb) -> dict:
    all_sheet_names = list(wb.sheetnames)
    sheets = []
    for name in all_sheet_names[:MAX_SHEETS]:
        ws = wb[name]
        rows, unreadable, more_rows = _read_sheet(ws)
        sheets.append(
            {
                "name": name,
                "rows": rows,
                # Kept SEPARATE from truncated_rows. Folding them together
                # meant an empty sheet and a partially-read one produced the
                # same payload, so the viewer could only say "0 rows × 0
                # columns · more rows truncated" — which tells a judge nothing
                # about whether they are looking at an empty sheet, a parse
                # failure, or genuinely withheld data.
                "unreadable": unreadable,
                "truncated_rows": more_rows,
                "truncated_cols": ws.max_column is not None and ws.max_column > MAX_COLS,
            }
        )
    return {
        "kind": "sheets",
        "sheets": sheets,
        "truncated_sheets": len(all_sheet_names) > MAX_SHEETS,
    }


# (read_only, data_only) attempts, in order of preference.
#   read_only=True   fast, streaming, but brittle: it trusts the declared
#                    dimension and resolves values lazily, so a malformed file
#                    raises mid-iteration.
#   read_only=False  a different, eager parser that survives several
#                    malformations the streaming one does not.
#   data_only=False  last resort: show formula TEXT. An agent that writes
#                    formulas with a library that stores no cached result
#                    leaves data_only=True returning None for every such cell,
#                    i.e. a sheet that looks completely empty. The formula
#                    source is not the computed answer, but it is real content
#                    a judge can assess, and it beats a blank panel.
_XLSX_ATTEMPTS = ((True, True), (False, True), (False, False))


def _xlsx(data: bytes) -> dict:
    import openpyxl

    result = {"kind": "sheets", "sheets": [], "truncated_sheets": False}
    for read_only, data_only in _XLSX_ATTEMPTS:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=read_only, data_only=data_only)
        except Exception:
            continue
        try:
            attempt = _xlsx_from_workbook(wb)
        except Exception:
            continue
        finally:
            wb.close()
        result = attempt
        # Accept the first attempt that actually found content. `degraded`
        # tells the viewer that what it is showing came from a fallback path,
        # so it can say so rather than presenting formula text as if it were
        # computed values.
        if any(s["rows"] for s in attempt["sheets"]):
            if not data_only:
                result["degraded"] = "formula-text"
            elif not read_only:
                result["degraded"] = "full-parse"
            break
    return result


def _docx(data: bytes) -> dict:
    """Converts the document to real HTML (headings, lists, tables, bold/
    italic all preserved) rather than a flat text dump, so a judge sees the
    document's actual structure — that structure is part of what they're
    assessing. Images are dropped: they'd have to be inlined as data URIs
    and these documents can carry several MB of them."""
    import mammoth

    result = mammoth.convert_to_html(io.BytesIO(data), convert_image=mammoth.images.img_element(lambda _: {}))
    html, truncated = _clip(result.value)
    return {"kind": "rich_html", "html": html, "truncated": truncated}


def render_pptx_as_pdf(data: bytes) -> bytes:
    """Render a deck privately with LibreOffice Impress.

    The source and output live only in a temporary directory during
    conversion. Callers cache the resulting PDF alongside the original
    Mongo deliverable, so PowerPoint never has to be publicly reachable.
    """
    with tempfile.TemporaryDirectory(prefix="arena-pptx-") as directory:
        workdir = Path(directory)
        source = workdir / "presentation.pptx"
        profile = workdir / "lo-profile"
        source.write_bytes(data)
        try:
            subprocess.run(
                [
                    "soffice",
                    "--headless",
                    "--nologo",
                    "--nofirststartwizard",
                    f"-env:UserInstallation={profile.as_uri()}",
                    "--convert-to",
                    "pdf:impress_pdf_Export",
                    "--outdir",
                    str(workdir),
                    str(source),
                ],
                check=True,
                capture_output=True,
                timeout=PPTX_RENDER_TIMEOUT_SECONDS,
            )
        except FileNotFoundError as exc:
            raise RuntimeError("PowerPoint preview service is not installed") from exc
        except subprocess.TimeoutExpired as exc:
            raise RuntimeError("PowerPoint rendering timed out") from exc
        except subprocess.CalledProcessError as exc:
            raise RuntimeError("PowerPoint could not be rendered") from exc

        output = workdir / "presentation.pdf"
        if not output.is_file():
            raise RuntimeError("PowerPoint could not be rendered")
        pdf = output.read_bytes()
        if not pdf.startswith(b"%PDF") or len(pdf) > MAX_RENDERED_PDF_BYTES:
            raise RuntimeError("PowerPoint produced an invalid preview")
        return pdf


def _pdf_meta(data: bytes) -> dict:
    """Only the page count is read here; the browser's own PDF viewer
    renders the real bytes via a dedicated content endpoint — see
    FileViewer.jsx / the `inline=true` mode of the content route. Extracted
    text was deliberately dropped: these deliverables are laid-out
    documents (charts, tables, figures), and text extraction loses exactly
    the presentation a judge is assessing."""
    from pypdf import PdfReader

    try:
        total = len(PdfReader(io.BytesIO(data)).pages)
    except Exception:
        total = None
    return {"kind": "pdf", "total_pages": total}


def _text(data: bytes) -> dict:
    raw = data[: MAX_CHARS + 1].decode("utf-8", errors="replace")
    text, truncated = _clip(raw)
    return {"kind": "text", "text": text, "truncated": truncated}


def _markdown(data: bytes) -> dict:
    raw = data[: MAX_CHARS + 1].decode("utf-8", errors="replace")
    text, truncated = _clip(raw)
    return {"kind": "markdown", "text": text, "truncated": truncated}


def _html(data: bytes) -> dict:
    raw = data[: MAX_CHARS + 1].decode("utf-8", errors="replace")
    text, truncated = _clip(raw)
    return {"kind": "html", "html": text, "truncated": truncated}


def _json_file(data: bytes) -> dict:
    raw = data[: MAX_CHARS + 1].decode("utf-8", errors="replace")
    try:
        pretty = json.dumps(json.loads(raw), indent=2)
    except ValueError:
        pretty = raw  # not valid JSON — show it as-is rather than failing
    text, truncated = _clip(pretty)
    return {"kind": "text", "text": text, "truncated": truncated}


def _csv(data: bytes) -> dict:
    import csv as csv_mod

    text = data.decode("utf-8", errors="replace")
    rows = []
    for row in csv_mod.reader(io.StringIO(text)):
        if len(rows) >= MAX_ROWS:
            break
        rows.append([c[:MAX_CELL_CHARS] for c in row[:MAX_COLS]])
    return {"kind": "sheets", "sheets": [{"name": "sheet1", "rows": rows}]}


EXTRACTORS = {
    ".xlsx": _xlsx,
    ".xlsm": _xlsx,
    ".docx": _docx,
    ".pptx": lambda _data: {"kind": "pptx_pdf"},
    ".pdf": _pdf_meta,
    ".html": _html,
    ".htm": _html,
    ".json": _json_file,
    ".csv": _csv,
    ".txt": _text,
    ".md": _markdown,
    ".py": _text,
}


def build_preview(data: bytes | None, filename: str) -> dict:
    ext = os.path.splitext(filename)[1].lower()
    extractor = EXTRACTORS.get(ext)
    if extractor is None:
        return {"kind": "unavailable", "reason": f"No inline preview for {ext or 'this file type'}."}
    if not data:
        return {"kind": "unavailable", "reason": "This deliverable has no stored content."}
    try:
        return extractor(data)
    except Exception as exc:
        # An agent-produced file can be malformed in arbitrary ways; a bad
        # file must not take down the page that lists it.
        return {"kind": "unavailable", "reason": f"Could not read this file ({type(exc).__name__})."}
